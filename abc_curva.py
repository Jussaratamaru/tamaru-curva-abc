"""
abc_curva.py — Curva ABC hierárquica (layout Excel, 1 linha por registro).
"""

from __future__ import annotations

import html
import io
import re
from typing import Callable

import pandas as pd

_ABC_CSS = """
<style>
html, body {
  margin: 0; padding: 0; width: 100%; height: 100%;
  overflow: hidden; font-family: "Segoe UI", Tahoma, sans-serif;
}
.abc-table-shell {
  display: flex; align-items: stretch; width: 100%; height: calc(100% - 44px); gap: 0;
}
.abc-scroll-btn {
  flex: 0 0 30px; border: 1px solid #c5c9d0; background: #eef2f7; color: #1f2937;
  font-size: 16px; line-height: 1; cursor: pointer; padding: 0; z-index: 2;
}
.abc-scroll-btn:hover:not(:disabled) { background: #dde3ea; }
.abc-scroll-btn:disabled { opacity: 0.35; cursor: default; }
.abc-erp-wrap {
  flex: 1 1 auto; min-width: 0; overflow: auto;
  border-top: 1px solid #c5c9d0; border-bottom: 1px solid #c5c9d0;
  scrollbar-width: thin; scrollbar-color: #9ca3af #f3f4f6;
}
.abc-erp-wrap::-webkit-scrollbar { width: 10px; height: 12px; }
.abc-erp-wrap::-webkit-scrollbar-thumb { background: #9ca3af; border-radius: 6px; }
.abc-erp-wrap::-webkit-scrollbar-track { background: #f3f4f6; }
.abc-erp {
  border-collapse: collapse; width: max-content; min-width: 100%;
  font-family: "Segoe UI", Tahoma, sans-serif; font-size: 11px;
}
.abc-erp th { background: #d9dce3; color: #1a1a2e; font-weight: 600; padding: 6px 8px; border: 1px solid #c5c9d0; text-align: center; white-space: nowrap; }
.abc-erp th.grp { background: #b8bcc6; font-size: 10px; letter-spacing: 0.02em; }
.abc-erp th.grp-sep-abc, .abc-erp th.grp-sep-med,
.abc-erp thead tr:nth-child(2) th:nth-child(6),
.abc-erp thead tr:nth-child(2) th:nth-child(10),
.abc-erp tbody td:nth-child(6),
.abc-erp tbody td:nth-child(10) {
  border-right: 3px solid #6b7280 !important;
}
.abc-erp th.col-desc { min-width: 220px; max-width: 320px; text-align: left; }
.abc-erp td { padding: 5px 8px; border: 1px solid #e2e5ea; vertical-align: middle; }
.abc-erp td.num { text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }
.abc-erp td.abc-rank { text-align: center; font-weight: 600; width: 52px; min-width: 52px; }
.abc-erp td.nome {
  font-size: 11px; color: #111827; text-align: left;
  min-width: 220px; max-width: 320px; white-space: normal; word-break: break-word;
}
.abc-cl-a { background: #dcfce7 !important; }
.abc-cl-b { background: #fef9c3 !important; }
.abc-cl-c { background: #f3f4f6 !important; }
.abc-nivel-1 td { background: #dde3ea; font-weight: 700; }
.abc-nivel-1 td.nome { text-transform: uppercase; }
.abc-nivel-2 td { background: #eef2f7; font-weight: 700; }
.abc-folha td { background: #ffffff !important; font-weight: 400; }
.abc-folha td.abc-rank {
  background: #ffffff !important; font-style: italic; font-size: 9px;
  text-align: right; font-weight: 400; color: #4b5563;
}
.abc-folha td.nome {
  font-style: italic; font-size: 9px; text-align: right; font-weight: 400;
  text-transform: none; color: #374151; max-width: 320px;
}
.abc-row-hidden { display: none !important; }
.abc-toolbar {
  display: flex; align-items: center; justify-content: space-between; gap: 12px;
  margin: 0 0 6px 2px; flex-wrap: wrap;
}
.abc-hint { font-size: 10px; color: #6b7280; margin: 0; }
.abc-expand-all-lbl {
  font-size: 11px; color: #374151; cursor: pointer; user-select: none;
  display: flex; align-items: center; gap: 6px; white-space: nowrap;
}
.abc-expand-all-lbl input { cursor: pointer; width: 14px; height: 14px; margin: 0; }
.abc-expandable td.nome { cursor: pointer; user-select: none; }
.abc-tree-toggle {
  display: inline-block; width: 14px; color: #4b5563; font-size: 10px;
  margin-right: 3px; vertical-align: middle; flex-shrink: 0;
}
.abc-tree-toggle::before { content: "▸"; }
.abc-expandable.abc-open .abc-tree-toggle::before { content: "▾"; }
.abc-expandable:not(.abc-open) .abc-tree-toggle::before { content: "▸"; }
</style>
"""

_ABC_TREE_JS = """
<script>
(function () {
  function collapseDescendants(tr) {
    var id = tr.id.replace("abc-row-", "");
    tr.classList.remove("abc-open");
    document.querySelectorAll('tr[data-parent="' + id + '"]').forEach(function (child) {
      child.classList.add("abc-row-hidden");
      collapseDescendants(child);
    });
  }
  function toggleRow(tr) {
    if (tr.getAttribute("data-has-children") !== "1") return;
    var open = tr.classList.toggle("abc-open");
    var id = tr.id.replace("abc-row-", "");
    document.querySelectorAll('tr[data-parent="' + id + '"]').forEach(function (child) {
      if (open) {
        child.classList.remove("abc-row-hidden");
      } else {
        child.classList.add("abc-row-hidden");
        collapseDescendants(child);
      }
    });
  }
  document.querySelectorAll(".abc-erp tbody").forEach(function (tbody) {
    tbody.addEventListener("click", function (e) {
      var tr = e.target.closest("tr.abc-expandable");
      if (tr) toggleRow(tr);
    });
  });

  function expandAllRows() {
    document.querySelectorAll(".abc-erp tbody tr").forEach(function (tr) {
      tr.classList.remove("abc-row-hidden");
      if (tr.classList.contains("abc-expandable")) tr.classList.add("abc-open");
    });
  }
  function collapseToLevel1() {
    document.querySelectorAll(".abc-erp tbody tr").forEach(function (tr) {
      var nivel = parseInt(tr.getAttribute("data-nivel") || "1", 10);
      tr.classList.remove("abc-open");
      if (nivel > 1) tr.classList.add("abc-row-hidden");
    });
  }
  var expandCb = document.getElementById("abc-expand-all");
  if (expandCb) {
    expandCb.addEventListener("change", function () {
      if (expandCb.checked) expandAllRows();
      else collapseToLevel1();
    });
  }

  var wrap = document.getElementById("abc-scroll-wrap");
  var btnL = document.getElementById("abc-scroll-left");
  var btnR = document.getElementById("abc-scroll-right");
  if (!wrap || !btnL || !btnR) return;

  function updateScrollBtns() {
    var max = wrap.scrollWidth - wrap.clientWidth;
    btnL.disabled = wrap.scrollLeft <= 2;
    btnR.disabled = max <= 2 || wrap.scrollLeft >= max - 2;
  }
  function scrollBy(dx) {
    wrap.scrollBy({ left: dx, behavior: "smooth" });
    setTimeout(updateScrollBtns, 280);
  }
  btnL.addEventListener("click", function () { scrollBy(-320); });
  btnR.addEventListener("click", function () { scrollBy(320); });
  wrap.addEventListener("scroll", updateScrollBtns);
  window.addEventListener("resize", updateScrollBtns);
  updateScrollBtns();
})();
</script>
"""

N_COLS_ABC = 6  # Curva ABC, Valor, Acumulada, tot%, acum%, Descrição

ABC_INFORMATIVO = (
    "Até 80% - Classe A: priorize visitas, estoque e negociação. "
    "Relacionamento próximo e alto nivel de serviço.",
    "80-95% - Classe B: Foco máximo em manter e buscar crescimento. Estimular faturamento mensal.",
    "Restante - Classe C: Revisar esforço comercial (baixo retorno relativo). "
    "Priorize canais economicos de atendimento.",
)

NIVEL_OPCOES = ["Fornecedor", "Cliente", "Produto", "Representante"]

GRUPOS_FORNECEDOR = ["SUMITOMO", "QUAKER", "INSIZE", "TAMARU", "OUTROS"]

NIVEL_COLUNAS: dict[str, tuple[str, str]] = {
    "Fornecedor": ("Fornecedor_Grupo", "Fornecedor_Grupo"),
    "Cliente": ("Cod.Cliente", "Cliente"),
    "Produto": ("Item", "Descrição do Item"),
    "Representante": ("Representante", "Representante"),
}

def calcular_curva_abc(
    df: pd.DataFrame,
    group_col: str,
    value_col: str = "Vl. Total",
    limites: tuple = (80.0, 95.0),
) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=[group_col, "Metrica", "Pct", "Pct_Acum", "Classe"])

    agg = (
        df.groupby(group_col, dropna=False)[value_col]
        .sum()
        .reset_index()
        .sort_values(value_col, ascending=False)
        .rename(columns={value_col: "Metrica"})
    )
    total = agg["Metrica"].sum()
    if total <= 0:
        agg["Pct"] = 0.0
        agg["Pct_Acum"] = 0.0
        agg["Classe"] = "C"
        return agg

    agg["Pct"] = agg["Metrica"] / total * 100
    agg["Pct_Acum"] = agg["Pct"].cumsum()
    lim_a, lim_b = limites

    def _classe(pct_acum: float) -> str:
        if pct_acum <= lim_a:
            return "A"
        if pct_acum <= lim_b:
            return "B"
        return "C"

    agg["Classe"] = agg["Pct_Acum"].apply(_classe)
    return agg


def meses_ate_ref(ref_mes: str, n: int = 12) -> list[str]:
    p = pd.Period(ref_mes, freq="M")
    return [str(p - i) for i in range(n - 1, -1, -1)]


def build_abc_legacy_report(
    df: pd.DataFrame,
    *,
    entity_id_col: str,
    entity_name_col: str,
    value_col: str,
    ref_mes: str,
) -> tuple[pd.DataFrame, list[str]]:
    meses_12 = meses_ate_ref(ref_mes, 12)
    df12 = df[df["AnoMes_str"].isin(meses_12)].copy()
    if df12.empty:
        return pd.DataFrame(), meses_12

    name_map = df12.groupby(entity_id_col)[entity_name_col].agg(
        lambda s: s.mode().iloc[0] if not s.mode().empty else (s.dropna().iloc[0] if s.notna().any() else "")
    )

    mensal = (
        df12.groupby([entity_id_col, "AnoMes_str"], dropna=False)[value_col]
        .sum()
        .reset_index()
    )
    pivot = mensal.pivot_table(
        index=entity_id_col,
        columns="AnoMes_str",
        values=value_col,
        aggfunc="sum",
        fill_value=0,
    )
    for m in meses_12:
        if m not in pivot.columns:
            pivot[m] = 0.0
    pivot = pivot[meses_12]

    totais = pivot.sum(axis=1).reset_index(name="_total")
    abc_in = totais.rename(columns={"_total": value_col})
    abc = calcular_curva_abc(abc_in, entity_id_col, value_col)
    abc = abc.rename(columns={"Metrica": "Valor"})
    abc["Acumulada"] = abc["Valor"].cumsum()
    abc["tot%"] = abc["Pct"]
    abc["acum%"] = abc["Pct_Acum"]
    abc[entity_name_col] = abc[entity_id_col].map(name_map)
    abc["Rank"] = range(1, len(abc) + 1)
    abc["Curva ABC"] = abc["Rank"].astype(str) + " " + abc["Classe"]

    out = abc.copy()
    for m in meses_12:
        out[m] = out[entity_id_col].map(pivot[m])
    out["Média 12"] = out[entity_id_col].map(pivot[meses_12].mean(axis=1))
    out["Média 09"] = out[entity_id_col].map(pivot[meses_12[-9:]].mean(axis=1))
    out["Média 06"] = out[entity_id_col].map(pivot[meses_12[-6:]].mean(axis=1))
    out["Média 03"] = out[entity_id_col].map(pivot[meses_12[-3:]].mean(axis=1))

    mes_labels = {m: pd.Period(m, freq="M").strftime("%m/%Y") for m in meses_12}
    rename_map = {entity_id_col: "Código", "Classe": "ABC", **mes_labels}
    if entity_name_col != entity_id_col:
        rename_map[entity_name_col] = "Descrição"
    out = out.rename(columns=rename_map)
    if entity_name_col == entity_id_col:
        out["Descrição"] = out["Código"]
    return out, meses_12


def build_abc_hierarquico(
    df: pd.DataFrame,
    niveis: list[tuple[str, int]],
    *,
    value_col: str,
    ref_mes: str,
    classe_filt: list[str] | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """
    Curva ABC aninhada: cada nível calcula A/B/C dentro do pai.
    Ex.: Fornecedor → Cliente → Produto → Representante.
    """
    if not niveis or df.empty:
        return pd.DataFrame(), meses_ate_ref(ref_mes, 12)

    classe_filt = classe_filt or ["A", "B", "C"]
    rows: list[dict] = []
    meses_12: list[str] = []

    def _walk(df_slice: pd.DataFrame, depth: int) -> None:
        nonlocal meses_12
        if depth >= len(niveis) or df_slice.empty:
            return

        dim_name, top_n = niveis[depth]
        id_col, name_col = NIVEL_COLUNAS[dim_name]

        raw, meses = build_abc_legacy_report(
            df_slice,
            entity_id_col=id_col,
            entity_name_col=name_col,
            value_col=value_col,
            ref_mes=ref_mes,
        )
        if not meses_12:
            meses_12 = meses
        if raw.empty:
            return

        show = raw[raw["ABC"].isin(classe_filt)].head(int(top_n))
        for _, row in show.iterrows():
            r = row.to_dict()
            r["Nivel"] = depth + 1
            r["NivelNome"] = dim_name
            rows.append(r)
            eid = row["Código"]
            child = df_slice[df_slice[id_col].astype(str) == str(eid)]
            _walk(child, depth + 1)

    _walk(df.copy(), 0)
    if not rows:
        return pd.DataFrame(), meses_12 or meses_ate_ref(ref_mes, 12)
    return pd.DataFrame(rows), meses_12


def fmt_num_erp(value: float, *, decimals: int = 0) -> str:
    if pd.isna(value) or value == 0:
        return "—"
    if decimals == 0:
        s = f"{int(round(value)):,}".replace(",", ".")
    else:
        s = f"{value:,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return s


def fmt_pct_erp(value: float) -> str:
    if pd.isna(value):
        return "—"
    return f"{value:.2f}".replace(".", ",")


def _indent_nivel(nivel: int) -> int:
    return max(0, (nivel - 1) * 14)


def _estilo_linha(nivel: int, max_nivel: int) -> str:
    """folha = último nível (≥3 camadas); n1/n2 = cabeçalhos hierárquicos."""
    if max_nivel >= 3 and nivel == max_nivel:
        return "folha"
    if nivel == 1:
        return "n1"
    return "n2"


def _classe_tr(nivel: int, max_nivel: int) -> str:
    est = _estilo_linha(nivel, max_nivel)
    if est == "folha":
        return "abc-folha"
    return f"abc-nivel-{1 if est == 'n1' else 2}"


def _compute_tree_meta(df: pd.DataFrame) -> tuple[list[int | None], list[bool]]:
    """Para cada linha (ordem depth-first): índice do pai e se tem filhos."""
    n = len(df)
    parents: list[int | None] = [None] * n
    has_children = [False] * n
    stack: dict[int, int] = {}
    for i in range(n):
        nivel = int(df.iloc[i]["Nivel"])
        if nivel > 1:
            parents[i] = stack.get(nivel - 1)
        stack[nivel] = i
        for k in list(stack):
            if k > nivel:
                del stack[k]
    for i, p in enumerate(parents):
        if p is not None:
            has_children[p] = True
    return parents, has_children


def render_abc_html_table(
    df: pd.DataFrame,
    meses_12: list[str],
    *,
    max_nivel: int = 3,
    currency: bool = True,
    max_rows: int | None = None,
) -> str:
    """HTML estilo Excel: 1 linha por registro, árvore expansível por clique."""
    mes_cols = [pd.Period(m, freq="M").strftime("%m/%Y") for m in meses_12]
    dec = 2 if currency else 0
    fmt_v = lambda v: fmt_num_erp(v, decimals=dec)

    show = df.head(max_rows) if max_rows else df
    show = show.reset_index(drop=True)
    parents, has_children = _compute_tree_meta(show)
    n_abc = N_COLS_ABC
    n_med = 4
    n_mes = len(mes_cols)

    parts = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'>",
        _ABC_CSS,
        "</head><body>",
        '<div class="abc-toolbar">',
        '<span class="abc-hint">Clique numa linha com ▸ para abrir ou fechar · ◀ ▶ ou barra inferior rolam as colunas.</span>',
        '<label class="abc-expand-all-lbl"><input type="checkbox" id="abc-expand-all"> Visualizar tudo aberto</label>',
        '</div>',
        '<div class="abc-table-shell">',
        '<button type="button" class="abc-scroll-btn" id="abc-scroll-left" title="Rolar para esquerda">◀</button>',
        '<div class="abc-erp-wrap" id="abc-scroll-wrap">',
        '<table class="abc-erp"><thead>',
    ]
    parts.append(
        f'<tr><th class="grp grp-sep-abc" colspan="{n_abc}">Curva ABC</th>'
        f'<th class="grp grp-sep-med" colspan="{n_med}">Média de vendas últimos</th>'
        f'<th class="grp" colspan="{n_mes}">Resultado dos Últimos 12 Meses</th></tr>'
    )
    parts.append(
        "<tr>"
        "<th>Curva ABC</th><th>Valor</th><th>Acumulada</th><th>tot%</th><th>acum%</th><th class='col-desc'>Descrição</th>"
        "<th>12</th><th>09</th><th>06</th><th>03</th>"
        + "".join(f"<th>{html.escape(c)}</th>" for c in mes_cols)
        + "</tr></thead><tbody>"
    )

    for i in range(len(show)):
        row = show.iloc[i]
        cls = str(row.get("ABC", "C"))
        rank = html.escape(str(row.get("Curva ABC", "")))
        nome = html.escape(str(row.get("Descrição", "")))
        nivel = int(row.get("Nivel", 1))
        is_folha = _estilo_linha(nivel, max_nivel) == "folha"
        rank_cls = "abc-rank" if is_folha else f"abc-rank abc-cl-{cls.lower()}"
        pad_nome = 0 if is_folha else 8 + _indent_nivel(nivel)

        tr_classes = [_classe_tr(nivel, max_nivel)]
        if nivel > 1:
            tr_classes.append("abc-row-hidden")
        if has_children[i]:
            tr_classes.append("abc-expandable")

        parent_attr = ""
        if parents[i] is not None:
            parent_attr = f' data-parent="{parents[i]}"'

        parts.append(
            f'<tr id="abc-row-{i}" class="{" ".join(tr_classes)}"'
            f' data-nivel="{nivel}" data-has-children="{"1" if has_children[i] else "0"}"'
            f"{parent_attr}>"
        )
        parts.append(f'<td class="{rank_cls}">{rank}</td>')
        parts.append(f'<td class="num">{fmt_v(row["Valor"])}</td>')
        parts.append(f'<td class="num">{fmt_v(row["Acumulada"])}</td>')
        parts.append(f'<td class="num">{fmt_pct_erp(row["tot%"])}</td>')
        parts.append(f'<td class="num">{fmt_pct_erp(row["acum%"])}</td>')
        toggle = '<span class="abc-tree-toggle"></span> ' if has_children[i] else ""
        estilo_nome = f' style="padding-left:{pad_nome}px"' if pad_nome else ""
        parts.append(f'<td class="nome"{estilo_nome}>{toggle}{nome}</td>')
        for mc in ["Média 12", "Média 09", "Média 06", "Média 03"]:
            parts.append(f'<td class="num">{fmt_v(row.get(mc, 0))}</td>')
        for mc in mes_cols:
            parts.append(f'<td class="num">{fmt_v(row.get(mc, 0))}</td>')
        parts.append("</tr>")

    parts.append("</tbody></table></div>")
    parts.append(
        '<button type="button" class="abc-scroll-btn" id="abc-scroll-right" title="Rolar para direita">▶</button>'
    )
    parts.append("</div>")
    if max_rows and len(df) > max_rows:
        parts.append(
            f'<p style="font-size:0.8rem;color:#6b7280;margin-top:6px;">'
            f"Exibindo {max_rows} de {len(df)} registros. Exporte o Excel para ver todos.</p>"
        )
    parts.extend([_ABC_TREE_JS, "</body></html>"])
    return "".join(parts)


def render_abc_table_widget(
    df: pd.DataFrame,
    meses_12: list[str],
    *,
    max_nivel: int = 3,
    currency: bool = True,
) -> None:
    """Renderiza tabela interativa (iframe) no Streamlit."""
    import streamlit.components.v1 as components

    html_doc = render_abc_html_table(
        df, meses_12, max_nivel=max_nivel, currency=currency,
    )
    n_rows = int((df["Nivel"] == 1).sum())
    height = min(960, max(420, 195 + n_rows * 34))
    components.html(html_doc, height=height, scrolling=False)


def _preencher_aba_abc(
    ws,
    df_raw: pd.DataFrame,
    meses_12: list[str],
    *,
    titulo: str,
    num_fmt: str,
    max_nivel: int = 3,
) -> None:
    """Preenche uma worksheet openpyxl com o layout ERP da Curva ABC."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    mes_labels = [pd.Period(m, freq="M").strftime("%m/%Y") for m in meses_12]
    media_cols = ["Média 12", "Média 09", "Média 06", "Média 03"]
    n_abc, n_med, n_mes = N_COLS_ABC, 4, len(mes_labels)
    n_total = n_abc + n_med + n_mes

    def _xls_val(v):
        if pd.isna(v) or v == 0:
            return "—"
        return v

    header_fill = PatternFill("solid", fgColor="B8BCC6")
    header_font = Font(bold=True, color="1A1A2E", size=10)
    sub_fill = PatternFill("solid", fgColor="D9DCE3")
    sub_font = Font(bold=True, size=9)
    row_fills = {
        1: PatternFill("solid", fgColor="DDE3EA"),
        2: PatternFill("solid", fgColor="EEF2F7"),
    }
    row_fill_branco = PatternFill("solid", fgColor="FFFFFF")
    abc_fills = {
        "A": PatternFill("solid", fgColor="DCFCE7"),
        "B": PatternFill("solid", fgColor="FEF9C3"),
        "C": PatternFill("solid", fgColor="F3F4F6"),
    }
    align_dir = Alignment(horizontal="right")
    info_fill = PatternFill("solid", fgColor="F8FAFC")

    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_total)

    hdr_grp = 3
    hdr_sub = 4
    row_ptr = 5

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_total)
    c_info = ws.cell(2, 1, "\n".join(ABC_INFORMATIVO))
    c_info.font = Font(size=9, color="374151")
    c_info.alignment = Alignment(wrap_text=True, vertical="top")
    c_info.fill = info_fill
    ws.row_dimensions[2].height = 52

    ws.merge_cells(start_row=hdr_grp, start_column=1, end_row=hdr_grp, end_column=n_abc)
    ws.cell(hdr_grp, 1, "Curva ABC").fill = header_fill
    ws.cell(hdr_grp, 1).font = header_font
    ws.merge_cells(start_row=hdr_grp, start_column=n_abc + 1, end_row=hdr_grp, end_column=n_abc + n_med)
    ws.cell(hdr_grp, n_abc + 1, "Média de vendas últimos").fill = header_fill
    ws.cell(hdr_grp, n_abc + 1).font = header_font
    ws.merge_cells(
        start_row=hdr_grp, start_column=n_abc + n_med + 1, end_row=hdr_grp, end_column=n_total,
    )
    ws.cell(hdr_grp, n_abc + n_med + 1, "Resultado dos Últimos 12 Meses").fill = header_fill
    ws.cell(hdr_grp, n_abc + n_med + 1).font = header_font

    sub_headers = (
        ["Curva ABC", "Valor", "Acumulada", "tot%", "acum%", "Descrição"]
        + ["12", "09", "06", "03"]
        + mes_labels
    )
    for ci, h in enumerate(sub_headers, 1):
        c = ws.cell(hdr_sub, ci, h)
        c.fill = sub_fill
        c.font = sub_font
        c.alignment = Alignment(horizontal="center")

    for _, r in df_raw.iterrows():
        cls = str(r.get("ABC", "C"))
        nivel = int(r.get("Nivel", 1))
        is_folha = _estilo_linha(nivel, max_nivel) == "folha"
        if is_folha:
            row_fill = row_fill_branco
        elif nivel == 1:
            row_fill = row_fills[1]
        else:
            row_fill = row_fills[2]
        font_nome = Font(
            bold=(not is_folha),
            italic=is_folha,
            size=9 if is_folha else 10,
        )
        font_rank = Font(
            bold=(not is_folha),
            italic=is_folha,
            size=9 if is_folha else 10,
        )

        c_rank = ws.cell(row_ptr, 1, r.get("Curva ABC", ""))
        c_rank.font = font_rank
        c_rank.alignment = align_dir if is_folha else Alignment(horizontal="center")
        if is_folha:
            c_rank.fill = row_fill_branco
        else:
            c_rank.fill = abc_fills.get(cls, abc_fills["C"])

        for ci, key in enumerate(["Valor", "Acumulada"], start=2):
            val = _xls_val(r.get(key, 0))
            c = ws.cell(row_ptr, ci, val)
            c.fill = row_fill
            if val != "—":
                c.number_format = num_fmt
                c.alignment = align_dir

        for ci, key in enumerate(["tot%", "acum%"], start=4):
            c = ws.cell(row_ptr, ci, r.get(key, 0) / 100)
            c.number_format = "0.00%"
            c.fill = row_fill
            c.alignment = align_dir

        c_nome = ws.cell(row_ptr, 6, r.get("Descrição", ""))
        c_nome.font = font_nome
        c_nome.fill = row_fill
        c_nome.alignment = align_dir if is_folha else Alignment(horizontal="left", indent=max(0, nivel - 1))

        for mi, mc in enumerate(media_cols, n_abc + 1):
            val = _xls_val(r.get(mc, 0))
            c = ws.cell(row_ptr, mi, val)
            c.fill = row_fill
            if val != "—":
                c.number_format = num_fmt
            c.alignment = align_dir

        for mi, ml in enumerate(mes_labels, n_abc + n_med + 1):
            val = _xls_val(r.get(ml, 0))
            c = ws.cell(row_ptr, mi, val)
            c.fill = row_fill
            if val != "—":
                c.number_format = num_fmt
            c.alignment = align_dir

        row_ptr += 1

    for ci in range(1, n_total + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions["F"].width = 30
    ws.freeze_panes = "A5"
    _configurar_impressao_abc(ws, hdr_sub=hdr_sub, last_row=row_ptr - 1, n_total=n_total)


def _configurar_impressao_abc(ws, *, hdr_sub: int, last_row: int, n_total: int) -> None:
    """A4 paisagem, 1 página de largura, col. A + linha de cabeçalho repetem; grade fina G:J."""
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    col_media_ini = N_COLS_ABC + 1  # G
    col_media_fim = N_COLS_ABC + 4    # J
    thin = Side(style="thin", color="AAAAAA")

    for r in range(hdr_sub, last_row + 1):
        for c in range(col_media_ini, col_media_fim + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.print_area = f"A1:{get_column_letter(n_total)}{last_row}"
    ws.print_title_rows = f"{hdr_sub}:{hdr_sub}"
    ws.print_title_cols = "A:A"
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.45, bottom=0.45, header=0.2, footer=0.2)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def exportar_abc_legacy_excel(
    df_raw: pd.DataFrame,
    meses_12: list[str],
    *,
    titulo: str,
    num_fmt: str,
    max_nivel: int = 3,
    sheet_name: str = "Curva ABC",
) -> bytes:
    """Excel 1 linha por entidade, estilo por nível hierárquico."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(sheet_name[:31])
        if "Sheet" in writer.book.sheetnames:
            del writer.book["Sheet"]
        _preencher_aba_abc(ws, df_raw, meses_12, titulo=titulo, num_fmt=num_fmt, max_nivel=max_nivel)
    buf.seek(0)
    return buf.getvalue()


def exportar_abc_legacy_excel_multi(
    blocos: list[tuple[str, pd.DataFrame, str]],
    meses_12: list[str],
    *,
    num_fmt: str,
    max_nivel: int = 3,
) -> bytes:
    """Excel com várias abas (ex.: um nível 1 por aba)."""
    if not blocos:
        raise ValueError("Nenhum bloco para exportar")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for i, (sheet_name, df_raw, titulo) in enumerate(blocos):
            ws = writer.book.create_sheet(sheet_name[:31])
            if i == 0 and "Sheet" in writer.book.sheetnames:
                del writer.book["Sheet"]
            _preencher_aba_abc(
                ws, df_raw, meses_12, titulo=titulo, num_fmt=num_fmt, max_nivel=max_nivel,
            )
    buf.seek(0)
    return buf.getvalue()


def _split_por_nivel1(df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    """Divide o relatório hierárquico em blocos (um por linha de nível 1)."""
    if df.empty:
        return []
    blocks: list[tuple[str, pd.DataFrame]] = []
    chunk_rows: list[dict] = []
    chunk_name = ""
    for _, row in df.iterrows():
        if int(row["Nivel"]) == 1:
            if chunk_rows:
                blocks.append((chunk_name, pd.DataFrame(chunk_rows)))
            chunk_name = str(row.get("Descrição") or row.get("Código") or "Nível 1")
            chunk_rows = [row.to_dict()]
        else:
            chunk_rows.append(row.to_dict())
    if chunk_rows:
        blocks.append((chunk_name, pd.DataFrame(chunk_rows)))
    return blocks


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    s = re.sub(r"[\\/?*\[\]:]", "", str(name)).strip()
    s = (s[:31] if s else "Aba")
    base = s
    n = 1
    while s in used:
        tail = f" {n}"
        s = (base[: 31 - len(tail)] + tail).strip() or f"Aba {n}"
        n += 1
    used.add(s)
    return s


def exportar_abc_por_nivel1(
    df_raw: pd.DataFrame,
    meses_12: list[str],
    *,
    titulo_base: str,
    num_fmt: str,
    max_nivel: int = 3,
) -> bytes:
    """Excel com uma aba por entidade do nível 1 (ex.: cada representante)."""
    blocos: list[tuple[str, pd.DataFrame, str]] = []
    used: set[str] = set()
    for nome, sub in _split_por_nivel1(df_raw):
        sheet = _sanitize_sheet_name(nome, used)
        blocos.append((sheet, sub, f"{titulo_base} · {nome}"))
    return exportar_abc_legacy_excel_multi(
        blocos, meses_12, num_fmt=num_fmt, max_nivel=max_nivel,
    )


def render_pagina_curva_abc(
    st,
    *,
    vendas: pd.DataFrame,
    vdf: pd.DataFrame,
    fmt_brl: Callable,
    fmt_num: Callable,
    fmt_mes: Callable,
    caixa_insight: Callable,
) -> None:
    st.caption(
        "Defina **1 a 4 níveis** (ex.: Fornecedor → Cliente → Produto). "
        "Cada nível calcula ABC **dentro do pai** · clique **Criar relatório** para gerar."
    )

    with st.container(border=True):
        st.markdown('<p class="panel-label">Hierarquia</p>', unsafe_allow_html=True)
        h1, h2 = st.columns([1, 5])
        with h1:
            qtd_niveis = st.number_input(
                "Qtd. níveis",
                min_value=1,
                max_value=4,
                value=3,
                step=1,
                key="abc_qtd_niveis",
                help="Ex.: 3 = Fornecedor → Cliente → Produto",
            )
        with h2:
            st.markdown(
                "<p style='margin:1.75rem 0 0;color:#64748b;font-size:0.85rem;'>"
                "Monte os níveis da esquerda para a direita (pai → filho).</p>",
                unsafe_allow_html=True,
            )

        _defaults = [
            ("Fornecedor", 50),
            ("Cliente", 50),
            ("Produto", 50),
            ("Representante", 50),
        ]
        niveis_cfg: list[tuple[str, int]] = []
        n_cols = int(qtd_niveis)
        lvl_cols = st.columns(n_cols)
        for i in range(n_cols):
            def_dim, def_top = _defaults[i]
            with lvl_cols[i]:
                st.markdown(
                    f"<p style='margin:0 0 4px;font-size:0.8rem;font-weight:600;color:#374151;'>"
                    f"Nível {i + 1}</p>",
                    unsafe_allow_html=True,
                )
                dim_n = st.selectbox(
                    "Dimensão",
                    NIVEL_OPCOES,
                    index=NIVEL_OPCOES.index(def_dim),
                    key=f"abc_nivel_dim_{i}",
                    label_visibility="collapsed",
                )
                top_n = st.number_input(
                    "Top",
                    min_value=1,
                    max_value=500,
                    value=def_top,
                    step=5,
                    key=f"abc_nivel_top_{i}",
                )
                niveis_cfg.append((dim_n, int(top_n)))

        dupes = [d for d in dict.fromkeys(niveis_cfg) if sum(1 for x, _ in niveis_cfg if x == d) > 1]
        if dupes:
            st.info(f"Níveis repetidos ({', '.join(dupes)}): a hierarquia pode ficar redundante.")

    with st.container(border=True):
        st.markdown('<p class="panel-label">Parâmetros do relatório</p>', unsafe_allow_html=True)
        g1, g2, g3, g4 = st.columns(4)
        with g1:
            metrica_sel = st.selectbox(
                "Indexador (A / B / C)",
                ["Valor (R$)", "Quantidade (un.)"],
                index=1,
                key="abc_metrica",
            )
        with g2:
            ref_mes = st.selectbox(
                "Mês de referência (12 meses)",
                sorted(vendas["AnoMes_str"].dropna().unique()),
                index=len(sorted(vendas["AnoMes_str"].dropna().unique())) - 1,
                key="abc_ref_mes",
            )
        with g3:
            classe_filt = st.multiselect(
                "Classe ABC", ["A", "B", "C"], default=["A", "B", "C"], key="abc_classe",
            )
        with g4:
            rep_sel = st.multiselect(
                "Representante",
                sorted(vendas["Representante"].dropna().unique()),
                default=[],
                placeholder="Todos",
                key="abc_rep",
            )

        st.markdown(
            "<p style='margin:0.75rem 0 0.35rem;font-size:0.72rem;font-weight:700;"
            "letter-spacing:0.06em;text-transform:uppercase;color:#64748b;'>"
            "Filtros opcionais</p>",
            unsafe_allow_html=True,
        )
        f1, f2, f3 = st.columns(3)
        with f1:
            forn_sel = st.multiselect(
                "Fornecedor (grupo)",
                GRUPOS_FORNECEDOR,
                default=[],
                placeholder="Todos",
                key="abc_forn",
            )
        with f2:
            cli_sel = st.multiselect(
                "Cliente",
                sorted(vendas["Cliente"].dropna().unique()),
                default=[],
                placeholder="Todos",
                key="abc_cli",
            )
        with f3:
            item_map = (
                vendas[["Item", "Descrição do Item"]]
                .dropna(subset=["Item"])
                .drop_duplicates("Item")
                .set_index("Item")["Descrição do Item"]
                .to_dict()
            )
            item_sel = st.multiselect(
                "Produto (código)",
                sorted(item_map.keys(), key=str),
                default=[],
                placeholder="Todos",
                key="abc_item",
                format_func=lambda i: f"{i} — {str(item_map.get(i, ''))[:40]}",
            )

    value_col = "Vl. Total" if metrica_sel == "Valor (R$)" else "Quantidade"
    currency = value_col == "Vl. Total"
    num_fmt = "#,##0.00" if currency else "#,##0"
    fmt_val = fmt_brl if currency else lambda v: fmt_num(v, 0)

    base = vdf.copy()
    if forn_sel:
        base = base[base["Fornecedor_Grupo"].isin(forn_sel)]
    if cli_sel:
        base = base[base["Cliente"].isin(cli_sel)]
    if item_sel:
        base = base[base["Item"].isin(item_sel)]
    if rep_sel:
        base = base[base["Representante"].isin(rep_sel)]

    if base.empty:
        st.warning("Nenhum dado no período/filtros atuais — ajuste o período ou filtros.")

    hier = " → ".join(f"{d} (top {n})" for d, n in niveis_cfg)
    max_nivel = len(niveis_cfg)

    st.markdown("")
    a1, a2 = st.columns([3, 1])
    with a1:
        st.markdown(
            f"<p style='margin:0.5rem 0 0;color:#374151;font-size:0.9rem;'>"
            f"<strong>Hierarquia:</strong> {hier}</p>",
            unsafe_allow_html=True,
        )
    with a2:
        criar = st.button(
            "📊 Criar relatório",
            type="primary",
            use_container_width=True,
            key="abc_criar",
        )

    if criar:
        if base.empty:
            st.session_state.pop("abc_relatorio", None)
            st.warning("Nenhum dado para os filtros selecionados.")
        else:
            filtros_xls = []
            if forn_sel:
                filtros_xls.append(f"Forn:{len(forn_sel)}")
            if cli_sel:
                filtros_xls.append(f"Cli:{len(cli_sel)}")
            if item_sel:
                filtros_xls.append(f"Prod:{len(item_sel)}")
            if rep_sel:
                filtros_xls.append(f"Rep:{len(rep_sel)}")
            titulo_xls = (
                f"Curva ABC — {hier} · {metrica_sel} · ref. {fmt_mes(ref_mes)}"
                + (f" · {' '.join(filtros_xls)}" if filtros_xls else "")
            )
            suffix = "valor" if currency else "qtd"
            slug = "_".join(d.lower() for d, _ in niveis_cfg)
            xls_file = f"ABC_{suffix}_{slug}_{ref_mes}.xlsx"
            with st.spinner("Gerando Curva ABC e Excel… (pode levar alguns segundos)"):
                raw, meses = build_abc_hierarquico(
                    base,
                    niveis_cfg,
                    value_col=value_col,
                    ref_mes=ref_mes,
                    classe_filt=classe_filt,
                )
                xls_bytes = None
                xls_err = None
                if not raw.empty:
                    try:
                        xls_bytes = exportar_abc_por_nivel1(
                            raw,
                            meses,
                            titulo_base=titulo_xls,
                            num_fmt=num_fmt,
                            max_nivel=max_nivel,
                        )
                    except Exception as err:
                        xls_err = str(err)
            st.session_state["abc_relatorio"] = {
                "raw": raw,
                "meses": meses,
                "hier": hier,
                "max_nivel": max_nivel,
                "currency": currency,
                "num_fmt": num_fmt,
                "metrica_sel": metrica_sel,
                "ref_mes": ref_mes,
                "niveis_cfg": niveis_cfg,
                "forn_sel": forn_sel,
                "cli_sel": cli_sel,
                "item_sel": item_sel,
                "rep_sel": rep_sel,
                "xls": xls_bytes,
                "xls_file": xls_file,
                "xls_err": xls_err,
            }

    rel = st.session_state.get("abc_relatorio")
    if not rel:
        st.info("Configure os filtros acima e clique em **Criar relatório** para gerar a tabela.")
        return

    raw = rel["raw"]
    meses = rel["meses"]
    max_nivel = rel["max_nivel"]
    currency = rel["currency"]
    num_fmt = rel["num_fmt"]
    metrica_sel = rel["metrica_sel"]
    ref_mes = rel["ref_mes"]
    niveis_cfg = rel["niveis_cfg"]
    hier = rel["hier"]
    forn_sel = rel.get("forn_sel", [])
    cli_sel = rel.get("cli_sel", [])
    item_sel = rel.get("item_sel", [])
    rep_sel = rel.get("rep_sel", [])

    if raw.empty:
        st.warning("Nenhum registro na curva ABC para os níveis selecionados.")
        return

    if rel:
        st.caption("Exibindo último relatório gerado · alterou filtros? Clique **Criar relatório** de novo.")

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Linhas no relatório", len(raw))
    k2.metric("Classe A", int((raw["ABC"] == "A").sum()))
    k3.metric(
        f"{'Receita' if currency else 'Volume'} classe A",
        fmt_val(raw.loc[raw["ABC"] == "A", "Valor"].sum()),
    )
    k4.metric(f"{'Receita' if currency else 'Volume'} total (nível 1)", fmt_val(
        raw.loc[raw["Nivel"] == 1, "Valor"].sum()
    ))

    xls = rel.get("xls")
    xls_file = rel.get("xls_file", "curva_abc.xlsx")
    if xls:
        n_abas = int((raw["Nivel"] == 1).sum())
        st.download_button(
            "📥 Exportar Curva ABC (.xlsx)",
            data=xls,
            file_name=xls_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="abc_download",
            help=f"Uma aba por entidade do nível 1 ({n_abas} abas).",
        )
    elif rel.get("xls_err"):
        st.warning(f"Exportação indisponível: {rel['xls_err']}")
    else:
        st.info("Clique em **Criar relatório** de novo para gerar o arquivo Excel.")

    st.markdown(
        '<div class="section-divider"><h3>Relatório Curva ABC</h3><div class="divider-line"></div></div>',
        unsafe_allow_html=True,
    )

    caixa_insight(list(ABC_INFORMATIVO), titulo="Classificação ABC", tipo="info")

    render_abc_table_widget(raw, meses, max_nivel=max_nivel, currency=currency)
